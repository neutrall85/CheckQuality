import re
from typing import Iterator, Optional, List, Callable
import openpyxl
from core.interfaces.i_loader import IDataLoader
from core.models.document import Document
from utils.date_parser import parse_date
from utils.string_utils import normalize_fio
import logging

logger = logging.getLogger(__name__)


class ExcelLoader(IDataLoader):
    """Загрузчик данных из Excel-файлов с итеративной обработкой."""

    # Список приоритетных имён листов (точное совпадение без учёта регистра)
    PRIORITY_SHEETS = ['Журнал', 'Журнал ', 'Замечания', 'Лист1']

    HEADER_ROW_MAX = 20

    def __init__(self, normalize_types: Optional[Callable[[str], str]] = None,
                 log_callback: Optional[Callable] = None):
        self._normalize_types = normalize_types or (lambda x: x)
        self._log_callback = log_callback

    def _log(self, message: str, level: str = "INFO"):
        if self._log_callback:
            self._log_callback(message, level)
        else:
            if level == "ERROR":
                logger.error(message)
            elif level == "WARNING":
                logger.warning(message)
            else:
                logger.info(message)

    def load(self, file_path: str) -> Iterator[Document]:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = self._get_target_sheet(wb)
            if ws is None:
                raise ValueError(f"Не найден подходящий лист в файле {file_path}. "
                                 f"Ожидались: {', '.join(self.PRIORITY_SHEETS)}")
            start_row = self._find_data_start_row(ws)
            for row in ws.iter_rows(min_row=start_row, max_col=8, values_only=True):
                if self._is_tech_row(row):
                    continue
                doc = self._parse_row(row)
                if doc:
                    yield doc
        finally:
            wb.close()

    def _get_target_sheet(self, wb) -> Optional[openpyxl.worksheet.worksheet.Worksheet]:
        """Ищет лист строго по приоритетным именам: Журнал -> Замечания -> Лист1."""
        sheets = {sheet_name.lower(): sheet_name for sheet_name in wb.sheetnames}
        for priority_name in self.PRIORITY_SHEETS:
            target_lower = priority_name.lower()
            if target_lower in sheets:
                return wb[sheets[target_lower]]
        return None

    def _find_data_start_row(self, ws) -> int:
        header_row = None
        max_row = min(ws.max_row, self.HEADER_ROW_MAX)
        for row_idx in range(1, max_row + 1):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, 9)]
            if any(v is not None for v in row_values) and self._looks_like_header(row_values):
                header_row = row_idx
                break
        return (header_row + 1) if header_row else 1

    def _looks_like_header(self, row: List) -> bool:
        for idx in (4, 5, 6):
            val = row[idx]
            if val is not None and isinstance(val, (int, float)):
                return False
        for idx in (3, 7):
            val = row[idx]
            if parse_date(val) is not None:
                return False
        return True

    def _is_tech_row(self, row) -> bool:
        if all(v is None for v in row):
            return True
        val_a = row[0]
        if isinstance(val_a, str) and re.match(r'^\d{4}-\d{2}-\d{2}', val_a.strip()):
            # Если первый столбец похож на дату, но есть другие значимые данные,
            # то не считаем строку технической
            if any(v is not None for v in row[2:8]):
                return False
            return True
        return False

    @staticmethod
    def _extract_int(value) -> int:
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str):
            match = re.search(r'\d+', value)
            if match:
                return int(match.group())
        return 0

    @staticmethod
    def _clean_string(value) -> str:
        if not isinstance(value, str):
            return ''
        cleaned = re.sub(r'<[^>]+>', '', value)
        return cleaned.strip()

    def _parse_developers(self, raw: str) -> List[str]:
        if not raw:
            return []
        raw = self._clean_string(raw)
        raw = re.sub(r'\s+и\s+', ',', raw, flags=re.IGNORECASE)
        raw = re.sub(r'\s{2,}', ',', raw)
        parts = re.split(r'[;,]+', raw)
        developers = []
        for p in parts:
            p = p.strip()
            if p:
                dev = normalize_fio(p)
                if dev:
                    developers.append(dev)
        return developers

    def _parse_row(self, row) -> Optional[Document]:
        try:
            a, b, c, d, e, f, g, h = row
        except ValueError:
            return None

        # Если все столбцы C–H пусты — пропускаем
        if all(v is None for v in (c, d, e, f, g, h)):
            return None

        doc_type = self._clean_string(a) if a else ''
        doc_type = self._normalize_types(doc_type)
        number = self._clean_string(b) if b else ''
        developers = self._parse_developers(c)
        receipt_date = parse_date(d)
        check_date = parse_date(h)

        # Итоговая дата — более поздняя из двух
        if receipt_date and check_date:
            final_date = max(receipt_date, check_date)
        else:
            final_date = check_date or receipt_date

        # Если даты нет — предупреждение и пропуск строки
        if final_date is None:
            self._log(
                f"Пропуск строки без даты: номер '{number}', тип '{doc_type}'. "
                f"Дата поступления: {d}, Дата проверки: {h}",
                level="WARNING"
            )
            return None

        a4_count = self._extract_int(e)
        errors_cat1 = self._extract_int(f)
        errors_cat2 = self._extract_int(g)

        return Document(
            doc_type=doc_type,
            number=number,
            developers=developers,
            receipt_date=receipt_date,
            a4_count=a4_count,
            errors_cat1=errors_cat1,
            errors_cat2=errors_cat2,
            check_date=final_date
        )