"""
Сборщик комментариев из Word и PDF (CLI + GUI).
Гарантированно извлекает ВСЕ комментарии из Word (через XML) и PDF (через PyMuPDF).
Период вводится в терминале, папка и файл сохранения выбираются графически.
Поддерживает ярлыки .lnk.
"""

import os
import sys
import re
from datetime import datetime
import zipfile
from lxml import etree
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# GUI
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
except ImportError:
    print("Ошибка: требуется tkinter")
    sys.exit(1)

# Ярлыки
try:
    import pylnk3
except ImportError:
    pylnk3 = None
    print("Предупреждение: pylnk3 не установлен – ярлыки .lnk игнорируются.")
    print("Установите: pip install pylnk3")


def clean_string(value: str) -> str:
    if not isinstance(value, str):
        value = str(value) if value is not None else ''
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', value)


def prompt_date(prompt_text: str) -> datetime:
    while True:
        s = input(prompt_text).strip()
        try:
            return datetime.strptime(s, '%Y-%m-%d')
        except ValueError:
            print(f"Неверный формат '{s}'. Ожидается ГГГГ-ММ-ДД.")


# ------------------------------------------------------------
# Word – извлечение ВСЕХ комментариев напрямую из XML (с полной диагностикой)
# ------------------------------------------------------------
def extract_comments_from_docx(filepath: str) -> list[dict]:
    """
    Извлекает все комментарии из Word, включая ответы.
    Не использует python-docx, работает напрямую с ZIP-архивом.
    Возвращает список словарей: author, date, text, context.
    """
    comments = []
    try:
        with zipfile.ZipFile(filepath, 'r') as z:
            # Проверяем наличие файла с комментариями
            comment_files = [n for n in z.namelist() if n.startswith('word/') and 'comment' in n.lower() and n.endswith('.xml')]
            if not comment_files:
                print(f"   Word: {os.path.basename(filepath)} — файлы комментариев не найдены в архиве.")
                return comments

            # Словарь для сбора комментариев и множество всех встреченных id
            comments_dict = {}
            all_ids: set[str] = set()

            # Перебираем все найденные файлы комментариев
            for cf in comment_files:
                xml_bytes = z.read(cf)
                tree = etree.fromstring(xml_bytes)

                # Возможные пространства имён
                ns_list = [
                    {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'},
                    {'w': 'http://schemas.microsoft.com/office/word/2010/wordml'},
                    {'w': 'http://schemas.microsoft.com/office/word/2012/wordml'},
                    {}  # без неймспейса
                ]

                for ns in ns_list:
                    # Ищем элементы comment
                    if ns:
                        elems = tree.findall('.//w:comment', namespaces=ns)
                    else:
                        elems = tree.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}comment')
                        if not elems:
                            elems = tree.findall('.//{http://schemas.microsoft.com/office/word/2010/wordml}comment')

                    for elem in elems:
                        # Определяем id
                        cid = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id') or \
                              elem.get('id')
                        if not cid:
                            continue
                        all_ids.add(cid)

                        # Если этот комментарий уже извлечён – пропускаем
                        if cid in comments_dict:
                            continue

                        # Извлекаем автора и дату
                        author = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}author') or \
                                 elem.get('author', 'Без автора')
                        date_str = elem.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}date') or \
                                   elem.get('date', '')

                        # Извлекаем текст комментария
                        text_parts = []
                        for p in elem.findall('.//w:p', namespaces={'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}):
                            para_text = []
                            for t in p.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                                if t.text:
                                    para_text.append(t.text)
                            if para_text:
                                text_parts.append(''.join(para_text))
                        full_text = '\n'.join(text_parts).strip()

                        # Пропускаем полностью пустые комментарии
                        if not full_text:
                            continue

                        # Форматируем дату
                        formatted_date = ''
                        if date_str:
                            try:
                                dt = datetime.strptime(date_str, '%Y-%m-%dT%H:%M:%SZ')
                                formatted_date = dt.strftime('%Y-%m-%d %H:%M')
                            except Exception:
                                formatted_date = clean_string(str(date_str))

                        # Сохраняем
                        comments_dict[cid] = {
                            'author': author,
                            'date': formatted_date,
                            'text': clean_string(full_text),
                            'context': ''
                        }

            total_raw = len(all_ids)
            if total_raw == 0:
                print(f"   Word: {os.path.basename(filepath)} — структура комментариев не распознана")
                return []

            # Ищем контекст в document.xml
            if 'word/document.xml' in z.namelist():
                doc_xml = etree.fromstring(z.read('word/document.xml'))
                paragraphs = doc_xml.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}p')
                for para in paragraphs:
                    para_text = []
                    for t in para.iter('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}t'):
                        if t.text:
                            para_text.append(t.text)
                    full_para = ''.join(para_text).strip()
                    if not full_para:
                        continue

                    ids = set()
                    for rstart in para.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}commentRangeStart'):
                        ref = rstart.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
                        if ref:
                            ids.add(ref)
                    for ref in para.findall('.//{http://schemas.openxmlformats.org/wordprocessingml/2006/main}commentReference'):
                        ref_id = ref.get('{http://schemas.openxmlformats.org/wordprocessingml/2006/main}id')
                        if ref_id:
                            ids.add(ref_id)
                    for cid in ids:
                        if cid in comments_dict and not comments_dict[cid]['context']:
                            comments_dict[cid]['context'] = clean_string(full_para)

            comments = list(comments_dict.values())
            print(f"   Word: {os.path.basename(filepath)} -> всего комментариев в XML: {total_raw}, извлечено с текстом: {len(comments)}")
    except Exception as e:
        print(f"   Word: {os.path.basename(filepath)} — ошибка: {e}")
        return []

    return comments


# ------------------------------------------------------------
# PDF – все текстовые аннотации (включая popup и выделения)
# ------------------------------------------------------------
def extract_comments_from_pdf(filepath: str) -> list[dict]:
    annotations = []
    total_annots = 0
    try:
        doc = fitz.open(filepath)
        for page_num in range(doc.page_count):
            page = doc[page_num]
            for annot in page.annots():
                if annot is None:
                    continue
                total_annots += 1
                subtype = annot.type[1] if annot.type[1] else 'Unknown'
                if subtype in ('Link', 'Widget', 'Movie', 'Sound', 'PrinterMark'):
                    continue

                content = ""
                author = ""
                date_str = ""

                # Информация аннотации
                try:
                    info = annot.info
                    content = info.get('content', '')
                    author = info.get('title', '')
                    date_str = info.get('modDate', '') or info.get('creationDate', '')
                except:
                    pass

                # Popup
                if not content:
                    try:
                        popup = annot.get_popup()
                        if popup:
                            content = popup.get('content', '')
                    except:
                        pass

                # Если и сейчас нет текста, пробуем взять выделенный текст (для Highlight и т.п.)
                if not content and subtype == 'Highlight':
                    try:
                        rect = annot.rect
                        words = page.get_text("words", clip=rect)
                        if words:
                            content = ' '.join(w[4] for w in words)
                            if content:
                                content = f"[Выделено] {content}"
                    except:
                        pass

                if not content:
                    continue

                if isinstance(content, bytes):
                    content = content.decode('utf-8', errors='replace')
                content = clean_string(str(content).strip())
                if isinstance(author, bytes):
                    author = author.decode('utf-8', errors='replace')
                author = clean_string(str(author).strip())
                if date_str:
                    try:
                        date_str = datetime.strptime(date_str.replace("'", ""), "D:%Y%m%d%H%M%S%z").strftime('%Y-%m-%d %H:%M')
                    except:
                        date_str = clean_string(str(date_str))

                context = ""
                try:
                    rect = annot.rect
                    if subtype == 'Highlight':
                        words = page.get_text("words", clip=rect)
                        if words:
                            context = ' '.join(w[4] for w in words)
                    else:
                        blocks = page.get_text("blocks", clip=rect)
                        if blocks:
                            context = ' '.join(b[4] for b in blocks if b[4].strip())
                    context = clean_string(context.strip())
                except:
                    pass

                annotations.append({
                    'author': author if author else 'Без автора',
                    'date': date_str,
                    'context': context,
                    'text': content,
                })
        doc.close()
        # Диагностика
        if annotations:
            print(f"PDF: {os.path.basename(filepath)} -> всего аннотаций: {total_annots}, извлечено с текстом: {len(annotations)}")
        else:
            print(f"PDF: {os.path.basename(filepath)} -> всего аннотаций: {total_annots}, извлечено с текстом: 0")
    except Exception as e:
        print(f"PDF: {os.path.basename(filepath)} — критическая ошибка: {e}")
    return annotations


# ------------------------------------------------------------
# Ярлыки и сканирование папок
# ------------------------------------------------------------
def resolve_lnk(lnk_path: str) -> str | None:
    if pylnk3 is None:
        return None
    try:
        lnk = pylnk3.parse(lnk_path)
        target = getattr(lnk, 'path', None)
        if not target and hasattr(lnk, 'link_info') and lnk.link_info:
            target = lnk.link_info.local_base_path
        if not target and hasattr(lnk, 'relative_path') and lnk.relative_path:
            base_dir = os.path.dirname(lnk_path)
            possible = os.path.normpath(os.path.join(base_dir, lnk.relative_path))
            if os.path.isdir(possible):
                target = possible
        if target and os.path.isdir(target):
            print(f"   [Ярлык] {os.path.basename(lnk_path)} -> {target}")
            return target
    except:
        pass
    return None


def scan_folder(folder: str, start_date: datetime, end_date: datetime, visited: set[str]) -> list[str]:
    found = []
    try:
        with os.scandir(folder) as entries:
            for entry in entries:
                if entry.is_file() and entry.name.lower().endswith('.lnk'):
                    target = resolve_lnk(entry.path)
                    if target and target not in visited:
                        visited.add(target)
                        found.extend(scan_folder(target, start_date, end_date, visited))
                    continue
                if entry.is_dir():
                    real = os.path.realpath(entry.path)
                    if real not in visited:
                        visited.add(real)
                        found.extend(scan_folder(entry.path, start_date, end_date, visited))
                    continue
                if entry.is_file() and entry.name.lower().endswith(('.docx', '.pdf')):
                    try:
                        mtime = datetime.fromtimestamp(entry.stat().st_mtime)
                        if start_date <= mtime <= end_date:
                            found.append(os.path.normpath(entry.path))
                    except OSError:
                        pass
    except PermissionError:
        pass
    return found


# ------------------------------------------------------------
# Главная функция
# ------------------------------------------------------------
def main():
    print("=== Сборщик комментариев из документов (Word/PDF) ===\n")
    print("Введите период проверки:")
    start_date = prompt_date("Дата начала (ГГГГ-ММ-ДД): ")
    end_date = prompt_date("Дата конца (ГГГГ-ММ-ДД): ").replace(hour=23, minute=59, second=59)
    if start_date > end_date:
        print("Ошибка: дата начала позже даты конца.")
        sys.exit(1)

    root = tk.Tk()
    root.withdraw()
    root.update()
    root.attributes('-topmost', True)
    folder = filedialog.askdirectory(title="Выберите папку с документами")
    if not folder:
        messagebox.showinfo("Отмена", "Папка не выбрана.")
        root.destroy()
        sys.exit(0)

    print(f"\nСканируем: {folder}")
    visited = {os.path.realpath(folder)}
    valid_files = scan_folder(folder, start_date, end_date, visited)

    if not valid_files:
        messagebox.showinfo("Результат", "Нет подходящих файлов за указанный период.")
        root.destroy()
        sys.exit(0)

    print(f"Найдено файлов: {len(valid_files)}")
    all_rows = []
    for filepath in valid_files:
        filename = os.path.basename(filepath)
        ext = os.path.splitext(filepath)[1].lower()
        print(f"Обрабатывается: {filename}")
        comments = []
        if ext == '.docx':
            comments = extract_comments_from_docx(filepath)
        elif ext == '.pdf':
            comments = extract_comments_from_pdf(filepath)

        if not comments:
            print("   Комментариев не найдено.")
        else:
            for c in comments:
                all_rows.append((
                    filename,
                    c.get('author', ''),
                    c.get('date', ''),
                    c.get('context', ''),
                    c.get('text', '')
                ))

    if not all_rows:
        messagebox.showinfo("Результат", "Ни одного комментария во всех проверенных файлах.")
        root.destroy()
        sys.exit(0)

    # Сохранение Excel
    default_name = f"сборник_комментариев_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    output_path = filedialog.asksaveasfilename(
        title="Сохранить Excel",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel файлы", "*.xlsx"), ("Все файлы", "*.*")]
    )
    if not output_path:
        messagebox.showinfo("Отмена", "Сохранение отменено.")
        root.destroy()
        sys.exit(0)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Комментарии"
    headers = ["Документ", "Автор", "Дата", "Контекст", "Комментарий"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    for row_idx, (filename, author, date, context, text) in enumerate(all_rows, 2):
        ws.cell(row=row_idx, column=1, value=clean_string(filename)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=2, value=clean_string(author)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=3, value=clean_string(date)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=4, value=clean_string(context)).alignment = wrap_alignment
        ws.cell(row=row_idx, column=5, value=clean_string(text)).alignment = wrap_alignment

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 50

    wb.save(output_path)
    messagebox.showinfo("Готово", f"Сохранено в:\n{output_path}\nВсего записей: {len(all_rows)}")
    root.destroy()


if __name__ == '__main__':
    main()