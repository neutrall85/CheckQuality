# report_comparator.py
import sys, os, re
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any, Optional, Tuple

# ---------------------------------------------------------------------------
# СТИЛИ
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
GROUP_FONT = Font(bold=True, size=12)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# ---------------------------------------------------------------------------
# РУССКИЕ МЕСЯЦЫ
# ---------------------------------------------------------------------------
RU_MONTHS = [
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь'
]
RU_MONTH_MAP = {name.lower(): i for i, name in enumerate(RU_MONTHS, start=1)}

def parse_month_key(name: str) -> Tuple[int, int]:
    parts = name.strip().split()
    if len(parts) >= 2:
        month_str = parts[0].lower()
        year_str = parts[-1]
        month = RU_MONTH_MAP.get(month_str, 0)
        try:
            year = int(year_str)
        except ValueError:
            year = 9999
        return (year, month)
    return (9999, 0)

# ---------------------------------------------------------------------------
# ИЗВЛЕЧЕНИЕ ДАННЫХ ИЗ ЛИСТА «Сводный отчёт»
# ---------------------------------------------------------------------------
def _find_row_by_text(ws, text: str, start_row: int = 1) -> Optional[int]:
    for row in range(start_row, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and isinstance(cell_value, str) and text in cell_value:
            return row
    return None

def _is_empty_row(ws, row: int, max_col: int) -> bool:
    for col in range(1, max_col + 1):
        val = ws.cell(row=row, column=col).value
        if val is not None and str(val).strip() != '':
            return False
    return True

def _read_table(ws, headers_row: int, max_col: int) -> List[Dict[str, Any]]:
    headers = []
    for col in range(1, max_col + 1):
        h = ws.cell(row=headers_row, column=col).value
        headers.append(str(h).strip() if h else f'Col{col}')
    data = []
    for row in range(headers_row + 1, ws.max_row + 1):
        if _is_empty_row(ws, row, max_col):
            break
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and isinstance(first_cell, str) and first_cell.startswith('ИТОГО'):
            break
        row_data = {}
        for col in range(1, max_col + 1):
            row_data[headers[col - 1]] = ws.cell(row=row, column=col).value
        data.append(row_data)
    return data

def _read_total_row(ws, section_title: str, col_headers: List[str]) -> Optional[Dict[str, Any]]:
    start = _find_row_by_text(ws, section_title)
    if start is None:
        return None
    header_row = None
    for row in range(start + 1, ws.max_row + 1):
        if _is_empty_row(ws, row, len(col_headers)):
            continue
        current = [ws.cell(row=row, column=c).value for c in range(1, len(col_headers) + 1)]
        if all(isinstance(v, str) and h in v for h, v in zip(col_headers, current)):
            header_row = row
            break
    if header_row is None:
        return None
    for row in range(header_row + 1, ws.max_row + 1):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and isinstance(first_cell, str) and first_cell.startswith('ИТОГО'):
            total_data = {}
            for col in range(1, len(col_headers) + 1):
                hdr = ws.cell(row=header_row, column=col).value
                hdr = str(hdr).strip() if hdr else f'Col{col}'
                total_data[hdr] = ws.cell(row=row, column=col).value
            return total_data
    return None

def extract_section(ws, section_title: str, col_headers: List[str]) -> List[Dict[str, Any]]:
    start = _find_row_by_text(ws, section_title)
    if start is None:
        raise ValueError(f"Секция '{section_title}' не найдена")
    header_row = None
    for row in range(start + 1, ws.max_row + 1):
        if _is_empty_row(ws, row, len(col_headers)):
            continue
        current = [ws.cell(row=row, column=c).value for c in range(1, len(col_headers) + 1)]
        if all(isinstance(v, str) and h in v for h, v in zip(col_headers, current)):
            header_row = row
            break
    if header_row is None:
        raise ValueError(f"Не найдены заголовки в секции '{section_title}'")
    return _read_table(ws, header_row, len(col_headers))

def extract_period_label(ws) -> str:
    for row in range(1, min(ws.max_row + 1, 20)):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and isinstance(cell_value, str) and 'Период данных' in cell_value:
            if ':' in cell_value:
                return cell_value.split(':', 1)[1].strip()
            return cell_value
    return 'Период не указан'

def extract_period_dates(ws) -> Optional[Tuple[date, date]]:
    label = extract_period_label(ws)
    matches = re.findall(r'(\d{2}\.\d{2}\.\d{4})', label)
    if len(matches) >= 2:
        try:
            d1 = datetime.strptime(matches[0], '%d.%m.%Y').date()
            d2 = datetime.strptime(matches[1], '%d.%m.%Y').date()
            return (min(d1, d2), max(d1, d2))
        except ValueError:
            return None
    return None

# ---------------------------------------------------------------------------
# СРАВНЕНИЕ ТАБЛИЦ
# ---------------------------------------------------------------------------
def build_merged_table(table1: List[Dict], table2: List[Dict],
                       key_field: str,
                       numeric_fields: List[str],
                       match_position: bool = False,
                       output_key: Optional[str] = None) -> List[Dict]:
    if output_key is None:
        output_key = key_field

    def safe_key(obj, field):
        val = obj.get(field)
        return str(val) if val is not None else ''

    if match_position:
        t1_sorted = sorted(table1, key=lambda r: parse_month_key(safe_key(r, key_field)))
        t2_sorted = sorted(table2, key=lambda r: parse_month_key(safe_key(r, key_field)))
        max_len = max(len(t1_sorted), len(t2_sorted))
        merged = []
        for i in range(max_len):
            row1 = t1_sorted[i] if i < len(t1_sorted) else None
            row2 = t2_sorted[i] if i < len(t2_sorted) else None
            month1 = row1[key_field] if row1 else "—"
            month2 = row2[key_field] if row2 else "—"
            composed_key = f"{month1} ↔ {month2}"
            entry = {output_key: composed_key}
            if row1 is None:
                status = 'added'
            elif row2 is None:
                status = 'removed'
            else:
                changed = False
                for field in numeric_fields:
                    if (row1.get(field) or 0) != (row2.get(field) or 0):
                        changed = True
                        break
                status = 'changed' if changed else 'unchanged'
            entry['status'] = status
            for field in numeric_fields:
                old_val = row1.get(field) if row1 else None
                new_val = row2.get(field) if row2 else None
                old_num = old_val if isinstance(old_val, (int, float)) else 0
                new_num = new_val if isinstance(new_val, (int, float)) else 0
                diff = new_num - old_num
                entry[f'{field}_old'] = old_num if row1 else 0
                entry[f'{field}_new'] = new_num if row2 else 0
                entry[f'{field}_diff'] = diff
                if old_num == 0 and new_num == 0:
                    pct = 0.0
                elif old_num == 0:
                    pct = None
                else:
                    pct = round((new_num - old_num) / old_num * 100, 1)
                entry[f'{field}_pct'] = pct
            merged.append(entry)
        return merged
    else:
        map1 = {safe_key(row, key_field): row for row in table1}
        map2 = {safe_key(row, key_field): row for row in table2}
        all_keys = sorted(set(map1.keys()) | set(map2.keys()))
        merged = []
        for key in all_keys:
            row1 = map1.get(key)
            row2 = map2.get(key)
            entry = {output_key: key}
            if row1 is None:
                status = 'added'
            elif row2 is None:
                status = 'removed'
            else:
                changed = False
                for field in numeric_fields:
                    if (row1.get(field) or 0) != (row2.get(field) or 0):
                        changed = True
                        break
                status = 'changed' if changed else 'unchanged'
            entry['status'] = status
            for field in numeric_fields:
                old_val = row1.get(field) if row1 else None
                new_val = row2.get(field) if row2 else None
                old_num = old_val if isinstance(old_val, (int, float)) else 0
                new_num = new_val if isinstance(new_val, (int, float)) else 0
                diff = new_num - old_num
                entry[f'{field}_old'] = old_num if row1 else 0
                entry[f'{field}_new'] = new_num if row2 else 0
                entry[f'{field}_diff'] = diff
                if old_num == 0 and new_num == 0:
                    pct = 0.0
                elif old_num == 0:
                    pct = None
                else:
                    pct = round((new_num - old_num) / old_num * 100, 1)
                entry[f'{field}_pct'] = pct
            merged.append(entry)
        return merged

# ---------------------------------------------------------------------------
# Вспомогательная функция округления чисел для вывода
# ---------------------------------------------------------------------------
def _round_val(v):
    """Округляет float до 2 знаков, остальное возвращает как есть."""
    if isinstance(v, float):
        return round(v, 2)
    return v

# ---------------------------------------------------------------------------
# КОНФИГУРАЦИЯ СЕКЦИЙ
# ---------------------------------------------------------------------------
SECTIONS = {
    'types': {
        'title': 'СТАТИСТИКА ПО ТИПАМ ДОКУМЕНТОВ',
        'headers': ['Тип документа', 'Количество', 'Всего ошибок кат.1', 'Всего ошибок кат.2',
                    'Средн. кол-во ошибок кат.1 на док', 'Средн. кол-во ошибок кат.2 на док',
                    'Всего форматов А4', 'Средн. кол-во форматов А4',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4'],
        'key': 'Тип документа',
        'numeric': ['Количество', 'Всего ошибок кат.1', 'Всего ошибок кат.2',
                    'Средн. кол-во ошибок кат.1 на док', 'Средн. кол-во ошибок кат.2 на док',
                    'Всего форматов А4', 'Средн. кол-во форматов А4',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4'],
        'match_position': False,
        'total_key': 'ИТОГО'
    },
    'developers': {
        'title': 'СТАТИСТИКА ПО РАЗРАБОТЧИКАМ (ВСЕ АВТОРЫ)',
        'headers': ['Разработчик', 'Количество документов', 'Количество форматов А4',
                    'Всего ошибок кат.1', 'Всего ошибок кат.2',
                    'Средн. кол-во ошибок кат.1 на док', 'Средн. кол-во ошибок кат.2 на док',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4',
                    'Коэф. ошибок', 'Рейтинг'],
        'key': 'Разработчик',
        'numeric': ['Количество документов', 'Количество форматов А4',
                    'Всего ошибок кат.1', 'Всего ошибок кат.2',
                    'Средн. кол-во ошибок кат.1 на док', 'Средн. кол-во ошибок кат.2 на док',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4',
                    'Коэф. ошибок', 'Рейтинг'],
        'match_position': False,
        'total_key': 'ИТОГО'
    },
    'months': {
        'title': 'СТАТИСТИКА ПО МЕСЯЦАМ',
        'headers': ['Месяц', 'Количество документов', 'Ошибки кат.1', 'Ошибки кат.2',
                    'Всего А4', 'Средн. кол-во ошибок кат.1 на док',
                    'Средн. кол-во ошибок кат.2 на док',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4'],
        'key': 'Месяц',
        'output_key': 'Сравнение месяцев',
        'numeric': ['Количество документов', 'Ошибки кат.1', 'Ошибки кат.2',
                    'Всего А4', 'Средн. кол-во ошибок кат.1 на док',
                    'Средн. кол-во ошибок кат.2 на док',
                    'Средн. кол-во ошибок кат.1 на А4', 'Средн. кол-во ошибок кат.2 на А4'],
        'match_position': True,
        'total_key': 'ИТОГО'
    }
}

# ---------------------------------------------------------------------------
# СОХРАНЕНИЕ В EXCEL
# ---------------------------------------------------------------------------
def save_combined_comparison(results: Dict[str, List[Dict]],
                             totals: Dict[str, Dict[str, Any]],
                             early_label: str, late_label: str,
                             output_path: str) -> None:
    while True:
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Сравнение периодов"
            row = 1

            # =============== ИТОГОВЫЕ ПОКАЗАТЕЛИ ===============
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=25)
            cell = ws.cell(row=row, column=1, value='ИТОГОВЫЕ ПОКАЗАТЕЛИ')
            cell.font = Font(bold=True, size=14, color='1F4E79')
            row += 1

            total_anchors = {}   # запомним строку с итоговой строкой

            for section_key in ['types', 'developers', 'months']:
                cfg = SECTIONS[section_key]
                total_row = totals.get(section_key)
                if not total_row:
                    continue
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1+len(cfg['numeric'])*4)
                cell = ws.cell(row=row, column=1, value=cfg['title'] + ' — ИТОГО')
                cell.font = GROUP_FONT
                row += 1

                ws.cell(row=row, column=1, value='Показатель').font = HEADER_FONT
                ws.cell(row=row, column=1).fill = HEADER_FILL
                ws.cell(row=row, column=1).border = THIN_BORDER
                col = 2
                for field in cfg['numeric']:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
                    c = ws.cell(row=row, column=col, value=field)
                    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
                    for cc in range(col, col+4):
                        ws.cell(row=row, column=cc).border = THIN_BORDER
                        ws.cell(row=row, column=cc).fill = HEADER_FILL
                    col += 4
                row += 1

                ws.cell(row=row, column=1, value='').border = THIN_BORDER
                col = 2
                for field in cfg['numeric']:
                    sub_headers = [late_label, early_label, 'Изм.', '%']
                    for i, sh in enumerate(sub_headers):
                        c = ws.cell(row=row, column=col + i, value=sh)
                        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
                    col += 4
                row += 1

                ws.cell(row=row, column=1, value='').border = THIN_BORDER
                col = 2
                for field in cfg['numeric']:
                    old_val = _round_val(total_row.get(f'{field}_old', 0))
                    new_val = _round_val(total_row.get(f'{field}_new', 0))
                    diff_val = _round_val(total_row.get(f'{field}_diff', 0))
                    pct_val = total_row.get(f'{field}_pct', 0)
                    ws.cell(row=row, column=col, value=new_val).border = THIN_BORDER
                    ws.cell(row=row, column=col+1, value=old_val).border = THIN_BORDER
                    ws.cell(row=row, column=col+2, value=diff_val).border = THIN_BORDER
                    pct_cell = ws.cell(row=row, column=col+3)
                    if pct_val is None:
                        pct_cell.value = '—'
                    else:
                        pct_cell.value = pct_val
                        pct_cell.number_format = '0.0"%"'
                    pct_cell.border = THIN_BORDER
                    col += 4
                total_anchors[section_key] = row
                row += 2

            # =============== ДЕТАЛЬНЫЕ СЕКЦИИ ===============
            for section_key in ['types', 'developers', 'months']:
                cfg = SECTIONS[section_key]
                table = results[section_key]
                display_key = cfg.get('output_key', cfg['key'])
                num_cols = 1 + len(cfg['numeric']) * 4

                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
                cell = ws.cell(row=row, column=1, value=cfg['title'])
                cell.font = GROUP_FONT
                row += 1

                ws.cell(row=row, column=1, value=display_key).font = HEADER_FONT
                ws.cell(row=row, column=1).fill = HEADER_FILL
                ws.cell(row=row, column=1).border = THIN_BORDER
                col = 2
                for field in cfg['numeric']:
                    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+3)
                    c = ws.cell(row=row, column=col, value=field)
                    c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
                    for cc in range(col, col+4):
                        ws.cell(row=row, column=cc).border = THIN_BORDER
                        ws.cell(row=row, column=cc).fill = HEADER_FILL
                    col += 4
                row += 1

                ws.cell(row=row, column=1, value='').border = THIN_BORDER
                col = 2
                for field in cfg['numeric']:
                    sub_headers = [late_label, early_label, 'Изм.', '%']
                    for i, sh in enumerate(sub_headers):
                        c = ws.cell(row=row, column=col + i, value=sh)
                        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.border = THIN_BORDER
                    col += 4
                row += 1

                for entry in table:
                    ws.cell(row=row, column=1, value=entry[display_key]).border = THIN_BORDER
                    col = 2
                    for field in cfg['numeric']:
                        old_val = _round_val(entry[f'{field}_old'])
                        new_val = _round_val(entry[f'{field}_new'])
                        diff_val = _round_val(entry[f'{field}_diff'])
                        pct_val = entry[f'{field}_pct']
                        ws.cell(row=row, column=col, value=new_val).border = THIN_BORDER
                        ws.cell(row=row, column=col+1, value=old_val).border = THIN_BORDER
                        ws.cell(row=row, column=col+2, value=diff_val).border = THIN_BORDER
                        pct_cell = ws.cell(row=row, column=col+3)
                        if pct_val is None:
                            if entry['status'] == 'added':
                                pct_cell.value = 'Новое'
                            elif entry['status'] == 'removed':
                                pct_cell.value = 'Удалено'
                            else:
                                pct_cell.value = '—'
                        else:
                            pct_cell.value = pct_val
                            pct_cell.number_format = '0.0"%"'
                        pct_cell.border = THIN_BORDER
                        col += 4
                    if entry['status'] == 'added':
                        fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif entry['status'] == 'removed':
                        fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    elif entry['status'] == 'changed':
                        fill = PatternFill(start_color='FFEB84', end_color='FFEB84', fill_type='solid')
                    else:
                        fill = None
                    if fill:
                        for c in range(1, num_cols + 1):
                            ws.cell(row=row, column=c).fill = fill
                    row += 1
                row += 2

            # Автоподбор ширины
            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                for r in range(1, ws.max_row + 1):
                    cell = ws.cell(row=r, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 30)

            # ================================================================
            # ЛИСТ С НАТИВНЫМИ ДИАГРАММАМИ ДЛЯ ТИПОВ ДОКУМЕНТОВ (ИТОГО)
            # Данные копируются с основного листа и диаграммы строятся здесь
            # ================================================================
            if 'types' in totals and totals['types']:
                ws_charts = wb.create_sheet("Графики итогов (типы)")
                cfg = SECTIONS['types']
                total_row_num = total_anchors['types']
                STEP_ROW = 12           # отступ между диаграммами (строк)
                CHART_WIDTH = 14
                CHART_HEIGHT = 18

                for idx, field in enumerate(cfg['numeric']):
                    # Извлекаем данные из основного листа
                    base_col = 2 + idx * 4       # поздний период
                    early_col = base_col + 1     # ранний период
                    pct_col = base_col + 3       # процент

                    late_val = ws.cell(row=total_row_num, column=base_col).value
                    early_val = ws.cell(row=total_row_num, column=early_col).value
                    pct_val = ws.cell(row=total_row_num, column=pct_col).value

                    data_start_row = 1 + idx * STEP_ROW

                    # Копируем данные в формате "Период" / "Значение"
                    ws_charts.cell(row=data_start_row, column=1, value="Период")
                    ws_charts.cell(row=data_start_row, column=2, value="Значение")
                    ws_charts.cell(row=data_start_row+1, column=1, value=early_label)
                    ws_charts.cell(row=data_start_row+1, column=2, value=_round_val(early_val))
                    ws_charts.cell(row=data_start_row+2, column=1, value=late_label)
                    ws_charts.cell(row=data_start_row+2, column=2, value=_round_val(late_val))

                    chart = BarChart()
                    chart.type = "col"
                    chart.style = 10
                    if pct_val is not None and isinstance(pct_val, (int, float)):
                        chart.title = f"{field}\n(изм. {pct_val:+.1f}%)"
                    else:
                        chart.title = field
                    chart.y_axis.title = None                # убираем "Значение"
                    chart.y_axis.majorGridlines = None       # убираем линии сетки
                    chart.x_axis.title = None

                    # Одна серия, две точки данных
                    data = Reference(ws_charts, min_col=2, min_row=data_start_row, max_row=data_start_row+2)
                    cats = Reference(ws_charts, min_col=1, min_row=data_start_row+1, max_row=data_start_row+2)
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)

                    series = chart.series[0]
                    pt0 = DataPoint(idx=0)
                    pt0.graphicalProperties.solidFill = "4472C4"   # Синий (ранний)
                    pt1 = DataPoint(idx=1)
                    pt1.graphicalProperties.solidFill = "FF8000"   # Оранжевый (поздний)
                    series.dPt = [pt0, pt1]

                    series.dLbls = DataLabelList()
                    series.dLbls.showVal = True
                    series.dLbls.showSerName = False
                    series.dLbls.showCatName = False
                    series.dLbls.dLblPos = 'ctr'

                    chart.width = CHART_WIDTH
                    chart.height = CHART_HEIGHT
                    anchor_cell = f"A{data_start_row}"
                    ws_charts.add_chart(chart, anchor_cell)

            wb.save(output_path)
            print(f"Результаты сохранены в: {os.path.abspath(output_path)}")
            return
        except PermissionError:
            print(f"Ошибка: Нет прав на запись файла '{output_path}'.")
            if sys.stdin.isatty():
                new_path = input("Введите новый путь для сохранения (или Enter для отмены): ").strip().strip('"')
                if new_path:
                    output_path = new_path
                    continue
                else:
                    print("Сохранение отменено.")
                    return
            else:
                print("Сохранение не выполнено.")
                return

# ---------------------------------------------------------------------------
# КОНСОЛЬНЫЙ ВЫВОД
# ---------------------------------------------------------------------------
def _fmt_num(v):
    """Форматирует число для консольного вывода: float => с двумя знаками, int => как есть."""
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)

def print_combined_report(results: Dict[str, List[Dict]],
                         totals: Dict[str, Dict[str, Any]],
                         early_label: str, late_label: str):
    print("\n===== ИТОГОВЫЕ ПОКАЗАТЕЛИ =====")
    for section_key in ['types', 'developers', 'months']:
        cfg = SECTIONS[section_key]
        total_row = totals.get(section_key)
        if not total_row:
            continue
        print(f"\n--- {cfg['title']} (Итого) ---")
        key_width = 30
        col_width = 15
        header_line = f"{'Показатель':<{key_width}}"
        for field in cfg['numeric']:
            header_line += f" {field + ' (ран.)':>{col_width}} {field + ' (позд.)':>{col_width}} {'Изм.':>8} {'%':>8}"
        print(header_line)
        sub_line = f"{'':<{key_width}}"
        for _ in cfg['numeric']:
            sub_line += f" {late_label:>{col_width}} {early_label:>{col_width}} {'':>8} {'':>8}"
        print(sub_line)
        line = f"{'ИТОГО':<{key_width}}"
        for field in cfg['numeric']:
            old = _fmt_num(total_row[f'{field}_old'])
            new = _fmt_num(total_row[f'{field}_new'])
            diff = _fmt_num(total_row[f'{field}_diff'])
            pct = total_row[f'{field}_pct']
            pct_str = f"{pct}%" if isinstance(pct, (int, float)) else '—'
            line += f" {new:>{col_width}} {old:>{col_width}} {diff:>8} {pct_str:>8}"
        print(line)

    for section_key in ['types', 'developers', 'months']:
        cfg = SECTIONS[section_key]
        table = results[section_key]
        display_key = cfg.get('output_key', cfg['key'])
        print(f"\n===== {cfg['title']} =====")
        if not table:
            print("  Нет данных.")
            continue
        key_width = 30 if section_key == 'months' else 25
        col_width = 15
        header_line = f"{'Ключ':<{key_width}}"
        for field in cfg['numeric']:
            header_line += f" {field + ' (ран.)':>{col_width}} {field + ' (позд.)':>{col_width}} {'Изм.':>8} {'%':>8}"
        print(header_line)
        sub_line = f"{'':<{key_width}}"
        for _ in cfg['numeric']:
            sub_line += f" {late_label:>{col_width}} {early_label:>{col_width}} {'':>8} {'':>8}"
        print(sub_line)
        for entry in table:
            line = f"{entry[display_key]:<{key_width}}"
            for field in cfg['numeric']:
                old = _fmt_num(entry[f'{field}_old'])
                new = _fmt_num(entry[f'{field}_new'])
                diff = _fmt_num(entry[f'{field}_diff'])
                pct = entry[f'{field}_pct']
                if pct is None:
                    if entry['status'] == 'added':
                        pct_str = 'Новое'
                    elif entry['status'] == 'removed':
                        pct_str = 'Удал.'
                    else:
                        pct_str = '—'
                else:
                    pct_str = f"{pct}%"
                line += f" {new:>{col_width}} {old:>{col_width}} {diff:>8} {pct_str:>8}"
            print(line)

# ---------------------------------------------------------------------------
# ГЛАВНАЯ ЛОГИКА СРАВНЕНИЯ
# ---------------------------------------------------------------------------
def compare_reports(file1: str, file2: str) -> Tuple[Dict[str, List[Dict]], Dict[str, Dict[str, Any]], str, str]:
    wb1 = openpyxl.load_workbook(file1, data_only=True)
    wb2 = openpyxl.load_workbook(file2, data_only=True)
    ws1 = wb1['Сводный отчёт']
    ws2 = wb2['Сводный отчёт']

    period1 = extract_period_label(ws1)
    period2 = extract_period_label(ws2)

    results = {}
    totals = {}
    for key, cfg in SECTIONS.items():
        t1 = extract_section(ws1, cfg['title'], cfg['headers'])
        t2 = extract_section(ws2, cfg['title'], cfg['headers'])
        merged = build_merged_table(t1, t2,
                                    key_field=cfg['key'],
                                    numeric_fields=cfg['numeric'],
                                    match_position=cfg.get('match_position', False),
                                    output_key=cfg.get('output_key'))
        results[key] = merged

        total1 = _read_total_row(ws1, cfg['title'], cfg['headers'])
        total2 = _read_total_row(ws2, cfg['title'], cfg['headers'])
        if total1 and total2:
            total1['ИТОГО'] = 'ИТОГО'
            total2['ИТОГО'] = 'ИТОГО'
            merged_total = build_merged_table([total1], [total2],
                                              key_field='ИТОГО',
                                              numeric_fields=cfg['numeric'])
            if merged_total:
                totals[key] = merged_total[0]
        else:
            totals[key] = {}

    wb1.close()
    wb2.close()
    return results, totals, period1, period2

# ---------------------------------------------------------------------------
# ВЫБОР ФАЙЛОВ
# ---------------------------------------------------------------------------
def choose_files_via_gui() -> Optional[List[str]]:
    try:
        import tkinter as tk
        from tkinter import filedialog
    except ImportError:
        return None
    root = tk.Tk()
    root.withdraw()
    file_paths = filedialog.askopenfilenames(
        title="Выберите файлы отчётов (порядок не важен)",
        filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
    )
    root.destroy()
    if len(file_paths) >= 2:
        return list(file_paths)
    return None

def choose_files_console() -> Optional[List[str]]:
    print("Введите пути к файлам (минимум 2). Пустая строка — завершить ввод.")
    files = []
    while True:
        path = input(f"Файл {len(files)+1}: ").strip().strip('"')
        if not path:
            break
        if os.path.exists(path):
            files.append(path)
        else:
            print(f"Файл не найден: {path}")
    return files if len(files) >= 2 else None

# ---------------------------------------------------------------------------
# ТОЧКА ВХОДА
# ---------------------------------------------------------------------------
def main():
    args = sys.argv[1:]
    output_path = None
    input_files = []

    i = 0
    while i < len(args):
        if args[i] in ('-o', '--output'):
            i += 1
            if i < len(args):
                output_path = args[i]
        else:
            input_files.append(args[i])
        i += 1

    if len(input_files) < 2:
        print("Интерактивный выбор файлов...")
        files = choose_files_via_gui() or choose_files_console()
        if files:
            input_files = files
        else:
            print("Недостаточно файлов для сравнения.")
            return

    f1, f2 = input_files[0], input_files[1]

    # Определяем хронологический порядок
    ordered = False
    try:
        wb1 = openpyxl.load_workbook(f1, data_only=True)
        wb2 = openpyxl.load_workbook(f2, data_only=True)
        dates1 = extract_period_dates(wb1['Сводный отчёт'])
        dates2 = extract_period_dates(wb2['Сводный отчёт'])
        wb1.close()
        wb2.close()
        if dates1 and dates2:
            if dates1[0] > dates2[0] or (dates1[0] == dates2[0] and dates1[1] > dates2[1]):
                f1, f2 = f2, f1
            ordered = True
            print(f"Ранний период: {os.path.basename(f1)} ({extract_period_label(openpyxl.load_workbook(f1, data_only=True)['Сводный отчёт'])})")
            print(f"Поздний период: {os.path.basename(f2)} ({extract_period_label(openpyxl.load_workbook(f2, data_only=True)['Сводный отчёт'])})")
    except Exception as e:
        print(f"Предупреждение: не удалось определить даты периодов ({e}). Будет использован выбранный порядок.")

    if not ordered:
        print("Порядок файлов: первый считается ранним, второй – поздним (если даты не извлечены).")

    try:
        results, totals, early_label, late_label = compare_reports(f1, f2)
    except Exception as e:
        print(f"Ошибка при сравнении: {e}")
        return

    print_combined_report(results, totals, early_label, late_label)

    if not output_path:
        base_name = os.path.splitext(os.path.basename(f1))[0]
        comp_name = os.path.splitext(os.path.basename(f2))[0]
        output_path = f"comparison_{base_name}_vs_{comp_name}.xlsx"
    save_combined_comparison(results, totals, early_label, late_label, output_path)

if __name__ == "__main__":
    main()